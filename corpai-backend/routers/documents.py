"""
CorpAI — Router de Documentos.
Upload, listagem e deleção de documentos por namespace.
Formatos aceitos: PDF, DOCX, XLSX, TXT, MD
"""

import logging
import uuid
from typing import List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from pydantic import BaseModel

from config import settings
from middleware.auth import get_current_user, require_role, TokenData
from services.chroma import chroma_service
from services.ollama import ollama_service
from services.rag import chunkear_texto

logger = logging.getLogger(__name__)
router = APIRouter()

# Extensões permitidas
EXTENSOES_PERMITIDAS = {".pdf", ".docx", ".xlsx", ".txt", ".md"}


# ─── Schemas ─────────────────────────────────────────────────
class DocumentoResponse(BaseModel):
    document_id: str
    nome_arquivo: str
    total_chunks: int
    criado_em: str


class UploadResponse(BaseModel):
    mensagem: str
    document_id: str
    nome_arquivo: str
    total_chunks: int


# ─── Funções de extração de texto ────────────────────────────
async def extrair_texto_pdf(conteudo: bytes) -> str:
    """Extrai texto de um arquivo PDF."""
    from PyPDF2 import PdfReader
    import io

    reader = PdfReader(io.BytesIO(conteudo))
    texto = ""
    for page in reader.pages:
        texto_pagina = page.extract_text()
        if texto_pagina:
            texto += texto_pagina + "\n"
    return texto


async def extrair_texto_docx(conteudo: bytes) -> str:
    """Extrai texto de um arquivo DOCX."""
    from docx import Document
    import io

    doc = Document(io.BytesIO(conteudo))
    texto = ""
    for paragrafo in doc.paragraphs:
        if paragrafo.text:
            texto += paragrafo.text + "\n"
    return texto


async def extrair_texto_xlsx(conteudo: bytes) -> str:
    """Extrai texto de um arquivo XLSX."""
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(conteudo), read_only=True)
    texto = ""
    for sheet in wb.worksheets:
        texto += f"## Planilha: {sheet.title}\n"
        for row in sheet.iter_rows(values_only=True):
            valores = [str(cell) if cell is not None else "" for cell in row]
            texto += " | ".join(valores) + "\n"
    return texto


async def extrair_texto_txt(conteudo: bytes) -> str:
    """Extrai texto de um arquivo TXT ou MD."""
    try:
        return conteudo.decode("utf-8")
    except UnicodeDecodeError:
        return conteudo.decode("latin-1")


async def extrair_texto(nome_arquivo: str, conteudo: bytes) -> str:
    """Extrai texto de um arquivo baseado na extensão."""
    extensao = nome_arquivo.lower().rsplit(".", 1)[-1] if "." in nome_arquivo else ""

    extractors = {
        "pdf": extrair_texto_pdf,
        "docx": extrair_texto_docx,
        "xlsx": extrair_texto_xlsx,
        "txt": extrair_texto_txt,
        "md": extrair_texto_txt,
    }

    extractor = extractors.get(extensao)
    if extractor is None:
        raise ValueError(f"Formato não suportado: .{extensao}")

    return await extractor(conteudo)


# ─── Endpoints ───────────────────────────────────────────────
@router.post(
    "/upload",
    response_model=UploadResponse,
    summary="Upload de documento",
)
async def upload_documento(
    file: UploadFile = File(...),
    current_user: TokenData = Depends(require_role(["lider_setor", "admin"])),
):
    """
    Faz upload de um documento, extrai texto, chunkeia,
    gera embeddings e armazena no namespace do setor do usuário.
    """
    # Validar extensão
    nome = file.filename or "documento"
    extensao = "." + nome.lower().rsplit(".", 1)[-1] if "." in nome else ""

    if extensao not in EXTENSOES_PERMITIDAS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Formato não suportado. Formatos aceitos: {', '.join(EXTENSOES_PERMITIDAS)}",
        )

    # Ler conteúdo do arquivo
    conteudo = await file.read()

    if len(conteudo) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="O arquivo está vazio.",
        )

    # Extrair texto
    try:
        texto = await extrair_texto(nome, conteudo)
    except Exception as e:
        logger.error(
            "Erro ao extrair texto do documento.",
            extra={"arquivo": nome, "erro": str(e)},
        )
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Erro ao processar o documento: {str(e)}",
        )

    if not texto.strip():
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Não foi possível extrair texto do documento.",
        )

    # Chunkear
    chunks = chunkear_texto(texto)

    if not chunks:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="O documento não gerou nenhum chunk de texto válido.",
        )

    # Gerar embeddings
    try:
        embeddings = await ollama_service.generate_embeddings_batch(chunks)
    except Exception as e:
        logger.error(
            "Erro ao gerar embeddings.",
            extra={"arquivo": nome, "erro": str(e)},
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao gerar embeddings. Verifique se o Ollama e o modelo nomic-embed-text estão disponíveis.",
        )

    # Armazenar no ChromaDB
    document_id = str(uuid.uuid4())
    from datetime import datetime

    ids = [f"{document_id}_chunk_{i}" for i in range(len(chunks))]
    metadatas = [
        {
            "document_id": document_id,
            "nome_arquivo": nome,
            "chunk_index": i,
            "total_chunks": len(chunks),
            "setor": current_user.setor,
            "usuario": current_user.username,
            "criado_em": datetime.utcnow().isoformat(),
        }
        for i in range(len(chunks))
    ]

    try:
        chroma_service.add_documents(
            namespace=current_user.setor,
            documents=chunks,
            embeddings=embeddings,
            metadatas=metadatas,
            ids=ids,
        )
    except Exception as e:
        logger.error(
            "Erro ao armazenar no ChromaDB.",
            extra={"arquivo": nome, "erro": str(e)},
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao armazenar documento no banco vetorial.",
        )

    logger.info(
        "Documento processado e indexado com sucesso.",
        extra={
            "document_id": document_id,
            "arquivo": nome,
            "setor": current_user.setor,
            "chunks": len(chunks),
        },
    )

    return UploadResponse(
        mensagem="Documento processado e indexado com sucesso.",
        document_id=document_id,
        nome_arquivo=nome,
        total_chunks=len(chunks),
    )


@router.get(
    "/",
    response_model=List[DocumentoResponse],
    summary="Listar documentos do setor",
)
async def listar_documentos(
    current_user: TokenData = Depends(require_role(["lider_setor", "admin"])),
):
    """Lista todos os documentos indexados no namespace do setor do usuário."""
    # Admin pode listar de qualquer setor, mas por padrão lista do seu
    setor = current_user.setor

    try:
        docs = chroma_service.list_documents(setor)
    except Exception as e:
        logger.error(
            "Erro ao listar documentos.",
            extra={"setor": setor, "erro": str(e)},
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao listar documentos do setor.",
        )

    return [
        DocumentoResponse(
            document_id=d["document_id"],
            nome_arquivo=d["nome_arquivo"],
            total_chunks=d["total_chunks"],
            criado_em=d.get("criado_em", ""),
        )
        for d in docs
    ]


@router.delete(
    "/{document_id}",
    summary="Deletar documento",
)
async def deletar_documento(
    document_id: str,
    current_user: TokenData = Depends(require_role(["lider_setor", "admin"])),
):
    """Remove um documento e todos os seus chunks do namespace do setor."""
    try:
        chroma_service.delete_document(
            namespace=current_user.setor,
            document_id=document_id,
        )
    except Exception as e:
        logger.error(
            "Erro ao deletar documento.",
            extra={"document_id": document_id, "erro": str(e)},
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao remover documento do banco vetorial.",
        )

    logger.info(
        "Documento removido.",
        extra={
            "document_id": document_id,
            "setor": current_user.setor,
            "usuario": current_user.username,
        },
    )

    return {"mensagem": "Documento removido com sucesso."}
